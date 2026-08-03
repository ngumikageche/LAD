from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Iterable

from openpyxl import Workbook, load_workbook


MAX_IMPORT_ROWS = 5000

SUPPORTED_EXTENSIONS = {"csv", "xlsx"}


def normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def normalize_lookup(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_people_upload(upload, preferred_sheet: str | None = None) -> list[dict[str, str]]:
    filename = str(upload.filename or "")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("Upload a CSV or XLSX workbook")

    if extension == "csv":
        try:
            text = upload.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV files must use UTF-8 encoding") from exc
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("The upload has no header row")
        raw_rows: Iterable[dict] = reader
    else:
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("The XLSX workbook could not be read") from exc
        worksheet = (
            workbook[preferred_sheet]
            if preferred_sheet and preferred_sheet in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        values = worksheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise ValueError("The workbook has no header row")
        raw_rows = (
            {clean_cell(header): value for header, value in zip(headers, row)}
            for row in values
        )

    rows: list[dict[str, str]] = []
    for raw in raw_rows:
        normalized = {
            normalize_header(header): clean_cell(value)
            for header, value in raw.items()
            if normalize_header(header)
        }
        if not any(normalized.values()):
            continue
        rows.append(normalized)
        if len(rows) > MAX_IMPORT_ROWS:
            raise ValueError(f"Imports are limited to {MAX_IMPORT_ROWS} data rows")
    if not rows:
        raise ValueError("The upload contains no data rows")
    return rows


def first_value(row: dict[str, str], *aliases: str) -> str:
    for alias in aliases:
        value = row.get(normalize_header(alias), "").strip()
        if value:
            return value
    return ""


# ── Conflict handling ────────────────────────────────────────────────────────
#
# An import never overwrites anything until the uploader has said so. Every
# import endpoint accepts `on_conflict`:
#
#   absent   the caller has not decided yet. If the file touches records that
#            already exist, the endpoint writes nothing and returns 409 with a
#            list of what would change, so the UI can ask.
#   "skip"   leave existing records untouched and import only the new rows.
#   "update" write the sheet's values over the existing records.
#
# Updates are field-by-field: a filled cell overwrites, a blank cell keeps
# whatever is already stored. That way a partial spreadsheet cannot silently
# wipe columns it does not carry.

CONFLICT_MODES = {"skip", "update"}


def resolve_conflict_mode(raw: str | None) -> str | None:
    """`skip` / `update`, or None when the uploader has not chosen yet."""
    value = str(raw or "").strip().lower()
    return value if value in CONFLICT_MODES else None


def apply_if_present(target, field: str, value) -> bool:
    """
    Overwrite `target.field` only when the sheet supplied a value.

    Returns True when something actually changed, so callers can tell a real
    update apart from a row that resolved to no change at all.
    """
    if value in (None, ""):
        return False
    if getattr(target, field, None) == value:
        return False
    setattr(target, field, value)
    return True


def conflict_response(conflicts: list[dict], total_rows: int, noun: str) -> tuple[dict, int]:
    """The 409 payload that asks the uploader how to treat existing records."""
    return {
        "error": (
            f"{len(conflicts)} of {total_rows} rows match {noun} that already exist. "
            "Choose whether to skip or update them."
        ),
        "needs_conflict_decision": True,
        "conflict_count": len(conflicts),
        "new_count": total_rows - len(conflicts),
        "total_rows": total_rows,
        "conflicts": conflicts,
    }, 409


def _format_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=worksheet.max_column).coordinate}"
    for column in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
        worksheet.column_dimensions[column[0].column_letter].width = width


def build_template(
    headers: list[str],
    sheet_name: str,
    example: list[str] | None = None,
    reference_sheets: dict[str, tuple[list[str], list[list[str]]]] | None = None,
) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    if example:
        worksheet.append(example)
    _format_worksheet(worksheet)
    for reference_name, (reference_headers, reference_rows) in (reference_sheets or {}).items():
        reference_sheet = workbook.create_sheet(reference_name)
        reference_sheet.append(reference_headers)
        for row in reference_rows:
            reference_sheet.append(row)
        _format_worksheet(reference_sheet)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
