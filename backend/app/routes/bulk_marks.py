from __future__ import annotations

import csv
import io
import json
import uuid

from flask import Blueprint, request, send_file, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..extensions import db
from ..models.assessment import Assessment
from ..models.course import Course
from ..models.department import Department
from ..models.score import Score
from ..models.score_evidence import ScoreEvidence
from ..models.student import Student
from ..models.student_subject import StudentSubject
from ..models.subject import Subject
from ..models.enrollment import Enrollment
from ..models.module import Module
from ..models.trainer import Trainer
from ..models.term import Term
from ..models.trainer_subject import TrainerSubject
from .permissions import _is_admin
from ..services.bulk_people_import import conflict_response, resolve_conflict_mode
from ..services.subject_enrollment import (
    link_student_to_subject,
    student_subject_link_exists,
)
from ..services.score_evidence import (
    EVIDENCE_UPLOAD_FOLDER,
    can_access_score_evidence,
    remove_score_evidence_files,
    save_score_evidence_files,
    usable_score_evidence_files,
)
from .permissions import get_current_user, require_permission

bp = Blueprint("bulk_marks", __name__, url_prefix="/scores/bulk-marks")


CAT_FORMULA_SCENARIOS = {
    "scenario_1": {
        "label": "Scenario 1: CAT 1 /30, CAT 2 /40, CAT 3 /30",
        "totals": [30, 40, 30],
    },
    "scenario_2": {
        "label": "Scenario 2: CAT 1 /100, CAT 2 /100, CAT 3 /100",
        "totals": [100, 100, 100],
    },
    "scenario_3": {
        "label": "Scenario 3: CAT 1 /80, CAT 2 /60, CAT 3 /100",
        "totals": [80, 60, 100],
    },
}


def _cell_text(value) -> str:
    """Return a stable text value for CSV and XLSX cells."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_marks_upload(upload) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    """Read marks rows from either CSV or the first worksheet of an XLSX file."""
    filename = str(upload.filename or "")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if extension == "csv":
        try:
            content = upload.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV files must use UTF-8 encoding") from exc
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            raise ValueError("CSV is empty or has no headers")
        fieldnames = [_cell_text(field).lower() for field in reader.fieldnames]
        rows = []
        for row_number, raw in enumerate(reader, start=2):
            row = {
                _cell_text(key).lower(): _cell_text(value)
                for key, value in raw.items()
                if key is not None
            }
            if any(row.values()):
                rows.append((row_number, row))
        return fieldnames, rows

    if extension == "xlsx":
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("The XLSX workbook could not be read") from exc
        worksheet = (
            workbook["Marks Upload"]
            if "Marks Upload" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        values = worksheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise ValueError("The workbook has no header row")
        fieldnames = [_cell_text(header).lower() for header in headers]
        rows = []
        for row_number, values_row in enumerate(values, start=2):
            row = {
                header: _cell_text(value)
                for header, value in zip(fieldnames, values_row)
                if header
            }
            if any(row.values()):
                rows.append((row_number, row))
        return fieldnames, rows

    raise ValueError("Upload a CSV or XLSX file")


def _sole_subject_of_module(module_id, cache: dict) -> "uuid.UUID | None":
    """
    The module's only subject, or None when it has none or more than one.

    Attributing a subject-less mark is only safe when there is a single
    candidate; with two subjects in a module the guess would put one class's
    marks on another class's report.
    """
    if module_id in cache:
        return cache[module_id]
    subject_ids = [
        row[0] for row in db.session.query(Subject.id).filter(
            Subject.module_id == module_id,
            Subject.deleted_at.is_(None),
        ).limit(2).all()
    ]
    resolved = subject_ids[0] if len(subject_ids) == 1 else None
    cache[module_id] = resolved
    return resolved


def _normalise_term(
    raw: str | None,
    assessment: Assessment | None,
    lookup: dict[str, str],
    row_number,
    errors: list[str],
) -> str | None:
    """
    The canonical term name for an uploaded row.

    Returns the matching term's own spelling when the label is recognised, so
    "term 1 2026" and "TERM 1 2026 " both become "Term 1 2026". An unrecognised
    label is replaced by the assessment's term and reported as a warning — the
    marks are still saved, but under a term the reports can actually find.
    """
    label = (raw or "").strip()
    assessment_term = assessment.term.name if assessment and assessment.term else None

    if not label:
        return assessment_term

    matched = lookup.get(label.lower())
    if matched:
        return matched

    if assessment_term:
        errors.append(
            f"Row {row_number}: term '{label}' matches no term, so it was saved "
            f"as '{assessment_term}'. Marks under an unknown term are hidden "
            f"from every report."
        )
        return assessment_term

    errors.append(
        f"Row {row_number}: term '{label}' matches no term on the system. "
        f"Create it under Academic Terms, or these marks will not appear on "
        f"any term-scoped report."
    )
    return label


def _grade(marks: float, total: float) -> str:
    pct = (marks / total * 100) if total else 0
    if pct >= 80: return "A"
    if pct >= 70: return "B"
    if pct >= 60: return "C"
    if pct >= 50: return "D"
    return "F"


def _resolve_student(value: str) -> Student | None:
    """Look up student by STU-code or registration_number."""
    if not value:
        return None
    s = db.session.query(Student).filter_by(code=value).first()
    if s:
        return s
    return db.session.query(Student).filter_by(registration_number=value).first()


def _resolve_subject(value: str) -> Subject | None:
    """Look up subject by SUB-code or UUID."""
    if not value:
        return None
    s = db.session.query(Subject).filter_by(code=value).first()
    if s:
        return s
    try:
        return db.session.get(Subject, uuid.UUID(value))
    except (ValueError, TypeError):
        return None


def _resolve_module(value: str) -> Module | None:
    """Look up a current module by MOD-code or UUID."""
    if not value:
        return None
    module = db.session.query(Module).filter_by(code=value).first()
    if module:
        return module
    try:
        return db.session.get(Module, uuid.UUID(value))
    except (ValueError, TypeError):
        return None


def _resolve_cat_formula(value: str | None) -> dict | None:
    key = str(value or "").strip().lower()
    return CAT_FORMULA_SCENARIOS.get(key)


def _calculate_cat_formula_marks(row: dict[str, str], scenario: dict, assessment: Assessment | None) -> tuple[float | None, dict | None, list[str]]:
    totals = scenario["totals"]
    labels = ["cat_1_marks", "cat_2_marks", "cat_3_marks"]
    component_marks = []
    errors = []

    for label, total in zip(labels, totals):
        value = row.get(label, "")
        if value == "":
            errors.append(f"{label} is required for the selected CAT formula")
            continue
        try:
            mark = float(value)
        except ValueError:
            errors.append(f"{label} must be a number")
            continue
        if mark < 0 or mark > total:
            errors.append(f"{label} must be 0–{total}")
        component_marks.append(mark)

    if errors or len(component_marks) != len(totals):
        return None, None, errors

    final_percentage = (sum(component_marks) / sum(totals) * 100) if sum(totals) else 0
    assessment_total = assessment.total_marks if assessment else 100
    marks = round(final_percentage / 100 * assessment_total, 2)
    return marks, {
        "scenario": next(
            (key for key, config in CAT_FORMULA_SCENARIOS.items() if config is scenario),
            None,
        ),
        "label": scenario["label"],
        "component_marks": component_marks,
        "component_total_marks": totals,
        "final_percentage": round(final_percentage, 2),
    }, []


def _resolve_assessment(value: str) -> Assessment | None:
    """Look up assessment by ASM-code or UUID."""
    if not value:
        return None
    a = db.session.query(Assessment).filter_by(code=value).first()
    if a:
        return a
    try:
        return db.session.get(Assessment, uuid.UUID(value))
    except (ValueError, TypeError):
        return None


def _uploader_trainer(user) -> Trainer | None:
    """Trainer profile of the uploader, or None for admins uploading school-wide."""
    if _is_admin(user):
        return None
    return db.session.query(Trainer).filter(Trainer.user_id == user.id).first()


def _trainer_subject_ids(trainer: Trainer) -> set[uuid.UUID]:
    rows = db.session.query(TrainerSubject.subject_id).filter(
        TrainerSubject.trainer_id == trainer.id
    ).all()
    return {row[0] for row in rows if row[0]}


def _selected_subject(user) -> tuple[Subject | None, dict | None, int | None]:
    """
    Resolve the subject the uploader picked for the whole batch.

    Accepts `subject_code` (SUB001) or `subject_id` (UUID or code) from the form
    body or query string. Returns (subject, error, status).
    """
    raw = (
        request.form.get("subject_code")
        or request.form.get("subject_id")
        or request.args.get("subject_code")
        or request.args.get("subject_id")
        or ""
    ).strip()
    if not raw:
        return None, None, None

    subject = _resolve_subject(raw)
    if not subject or subject.deleted_at:
        return None, {"error": f"Subject '{raw}' not found"}, 404

    trainer = _uploader_trainer(user)
    if trainer and subject.id not in _trainer_subject_ids(trainer):
        return None, {"error": "You are not assigned to this subject"}, 403
    return subject, None, None


@bp.get("/subjects")
def list_subjects():
    """Subjects the uploader may attach marks to, keyed by subject code."""
    user, error, status = require_permission("scores.create")
    if error:
        return error, status

    q = db.session.query(Subject).filter(Subject.deleted_at.is_(None))

    trainer = _uploader_trainer(user)
    if trainer:
        subject_ids = _trainer_subject_ids(trainer)
        if not subject_ids:
            return {"subjects": [], "total": 0, "scope": "trainer"}, 200
        q = q.filter(Subject.id.in_(subject_ids))

    search = (request.args.get("search") or "").strip().lower()
    items = []
    for subject in q.order_by(Subject.code.asc(), Subject.name.asc()).all():
        module = getattr(subject, "module", None)
        course = getattr(module, "course", None)
        if search and search not in f"{subject.code or ''} {subject.name}".lower():
            continue
        items.append({
            "id": str(subject.id),
            "code": subject.code,
            "name": subject.name,
            "module_id": str(subject.module_id) if subject.module_id else None,
            "module_code": module.code if module else None,
            "module_name": module.name if module else None,
            "course_id": str(course.id) if course else None,
            "course_name": course.name if course else None,
        })

    return {
        "subjects": items,
        "total": len(items),
        "scope": "trainer" if trainer else "all",
    }, 200


@bp.get("/assessments")
def list_assessments():
    user, error, status = require_permission("scores.create")
    if error:
        return error, status

    course_id = request.args.get("course_id")
    subject_ref = request.args.get("subject_code") or request.args.get("subject_id")
    q = db.session.query(Assessment)
    q = q.filter(Assessment.assessment_scope == "formative")
    if course_id:
        try:
            q = q.filter(Assessment.course_id == uuid.UUID(course_id))
        except (ValueError, TypeError):
            pass
    if subject_ref:
        subject = _resolve_subject(subject_ref.strip())
        module = getattr(subject, "module", None) if subject else None
        if module and module.course_id:
            q = q.filter(Assessment.course_id == module.course_id)
            q = q.filter(Assessment.module_id == module.id)
    items = [
        {
            "id": str(a.id),
            "code": a.code,
            "name": a.name,
            "assessment_type": a.assessment_type,
            "assessment_scope": a.assessment_scope,
            "total_marks": a.total_marks,
            "pass_marks": a.pass_marks,
            "course_id": str(a.course_id) if a.course_id else None,
            "course_name": a.course.name if a.course else None,
        }
        for a in q.order_by(Assessment.name.asc()).all()
    ]
    return {"assessments": items, "total": len(items)}, 200


@bp.get("/cat-formulas")
def list_cat_formulas():
    user, error, status = require_permission("scores.create")
    if error:
        return error, status

    return {
        "formulas": [
            {
                "key": key,
                "label": config["label"],
                "component_total_marks": config["totals"],
                "total_marks": sum(config["totals"]),
            }
            for key, config in CAT_FORMULA_SCENARIOS.items()
        ],
    }, 200


@bp.post("/assessments")
def create_assessment():
    """Create an internal formative assessment for an accessible subject."""
    user, error, status = require_permission("scores.create")
    if error:
        return error, status

    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or "").strip()
    subject_ref = str(payload.get("subject_code") or payload.get("subject_id") or "").strip()
    assessment_type = str(payload.get("assessment_type") or "test").strip().lower()

    if not name:
        return {"error": "Assessment name is required"}, 400
    if not subject_ref:
        return {"error": "Select a subject"}, 400

    subject = _resolve_subject(subject_ref)
    if not subject or subject.deleted_at:
        return {"error": "Subject not found"}, 404

    trainer = _uploader_trainer(user)
    if trainer and subject.id not in _trainer_subject_ids(trainer):
        return {"error": "You are not assigned to this subject"}, 403

    module = subject.module
    course = module.course if module else None
    if not module or not course:
        return {"error": "The selected subject is not linked to a course"}, 400

    try:
        total_marks = int(payload.get("total_marks"))
        pass_marks = int(payload.get("pass_marks"))
    except (TypeError, ValueError):
        return {"error": "Total marks and pass marks must be whole numbers"}, 400
    if total_marks <= 0:
        return {"error": "Total marks must be greater than zero"}, 400
    if pass_marks < 0 or pass_marks > total_marks:
        return {"error": "Pass marks must be between zero and total marks"}, 400

    assessment = Assessment(
        name=name,
        assessment_type=assessment_type or "test",
        assessment_scope="formative",
        total_marks=total_marks,
        pass_marks=pass_marks,
        course_id=course.id,
        module_id=module.id,
    )
    db.session.add(assessment)
    db.session.commit()

    return {
        "id": str(assessment.id),
        "code": assessment.code,
        "name": assessment.name,
        "assessment_type": assessment.assessment_type,
        "assessment_scope": assessment.assessment_scope,
        "total_marks": assessment.total_marks,
        "pass_marks": assessment.pass_marks,
        "course_id": str(course.id),
        "course_name": course.name,
        "module_id": str(module.id),
        "module_name": module.name,
        "subject_id": str(subject.id),
        "subject_code": subject.code,
    }, 201


# ── Preview / validate ────────────────────────────────────────────────────────

@bp.post("/preview")
def preview_bulk():
    """
    Parse an uploaded CSV or XLSX file and return a preview with validation results.
    Required columns:
        student_id  (STU001 code OR registration_number)
        marks_obtained, or cat_1_marks/cat_2_marks/cat_3_marks with cat_formula
        assessment_code  (ASM001 code)
    Optional:
        assessment_id  (legacy; UUID or ASM-code)
        module_code  (MOD001 code)
        module_id  (legacy; MOD001 code OR module UUID)
        subject_code  (SUB001 code)
        subject_id  (legacy; SUB001 code OR subject UUID)
        term
        feedback

    A `subject_code` form field selects one subject for the whole batch; rows
    without their own subject code inherit it.
    """
    user, error, status = require_permission("scores.create")
    if error:
        return error, status

    batch_subject, subject_error, subject_status = _selected_subject(user)
    if subject_error:
        return subject_error, subject_status

    file = request.files.get("file")
    if not file:
        return {"error": "No file provided"}, 400

    try:
        fieldnames, uploaded_rows = _read_marks_upload(file)
    except ValueError as exc:
        return {"error": str(exc)}, 400

    cat_formula = _resolve_cat_formula(request.form.get("cat_formula"))
    if request.form.get("cat_formula") and not cat_formula:
        return {"error": "Unknown CAT formula scenario"}, 400

    fields = set(fieldnames)
    # Accept either student_id (new) or registration_number (legacy)
    has_student = "student_id" in fields or "registration_number" in fields
    required = set() if cat_formula else {"marks_obtained"}
    missing = required - fields
    if not has_student:
        missing.add("student_id")
    if "assessment_code" not in fields and "assessment_id" not in fields:
        missing.add("assessment_code")
    if cat_formula:
        missing.update({"cat_1_marks", "cat_2_marks", "cat_3_marks"} - fields)
    if missing:
        return {"error": f"Missing required columns: {', '.join(sorted(missing))}"}, 400

    # Cache lookups
    student_cache: dict[str, Student | None] = {}
    assessment_cache: dict[str, Assessment | None] = {}
    module_cache: dict[str, Module | None] = {}
    subject_cache: dict[str, Subject | None] = {}

    rows = []
    for i, row in uploaded_rows:

        # Accept student_id (new) or registration_number (legacy)
        student_key = row.get("student_id") or row.get("registration_number", "")
        marks_str = row.get("marks_obtained", "")
        row_cat_formula = cat_formula or _resolve_cat_formula(row.get("cat_formula"))
        if row.get("cat_formula") and not row_cat_formula:
            row_cat_formula = None
        assessment_id_str = row.get("assessment_code") or row.get("assessment_id", "")
        module_str = row.get("module_code") or row.get("module_id", "")
        # New exports use the accurately named subject_code column. Keep
        # accepting subject_id so previously downloaded files still work.
        subject_str = row.get("subject_code") or row.get("subject_id", "")
        term = row.get("term", "") or None
        feedback = row.get("feedback", "") or None

        errors = []

        # Validate assessment
        assessment = None
        if assessment_id_str:
            if assessment_id_str not in assessment_cache:
                assessment_cache[assessment_id_str] = _resolve_assessment(assessment_id_str)
            assessment = assessment_cache.get(assessment_id_str)
            if not assessment:
                errors.append(f"Assessment '{assessment_id_str}' not found")
            elif assessment.assessment_scope != "formative":
                errors.append("Only internal formative assessments can be uploaded")
        else:
            errors.append("assessment_code is required")

        # Validate marks, or calculate them from CAT component columns.
        formula_details = None
        if row.get("cat_formula") and not _resolve_cat_formula(row.get("cat_formula")):
            marks = None
            errors.append("Unknown CAT formula scenario")
        elif row_cat_formula:
            marks, formula_details, formula_errors = _calculate_cat_formula_marks(row, row_cat_formula, assessment)
            errors.extend(formula_errors)
        else:
            try:
                marks = float(marks_str)
            except ValueError:
                marks = None
                errors.append("marks_obtained must be a number")

        # Validate marks range
        if marks is not None and assessment:
            if marks < 0 or marks > assessment.total_marks:
                errors.append(f"marks_obtained must be 0–{assessment.total_marks}")

        # Validate student
        student = None
        if student_key:
            if student_key not in student_cache:
                student_cache[student_key] = _resolve_student(student_key)
            student = student_cache[student_key]
            if not student:
                errors.append(f"Student '{student_key}' not found")
        else:
            errors.append("student_id is required")

        # Resolve the row module. New templates provide module_code; older files
        # can inherit it from the selected subject or assessment.
        module = None
        module_input_invalid = False
        if module_str:
            if module_str not in module_cache:
                module_cache[module_str] = _resolve_module(module_str)
            module = module_cache[module_str]
            if not module or module.deleted_at:
                errors.append(f"Module '{module_str}' not found")
                module = None
                module_input_invalid = True

        # Resolve subject — a row's own subject_id wins, otherwise the batch subject
        subject = None
        if subject_str:
            if subject_str not in subject_cache:
                subject_cache[subject_str] = _resolve_subject(subject_str)
            subject = subject_cache[subject_str]
            if not subject:
                errors.append(f"Subject '{subject_str}' not found")
        elif batch_subject:
            subject = batch_subject

        if not module and not module_input_invalid:
            module = subject.module if subject else (assessment.module if assessment else None)
        if module and assessment and assessment.module_id and assessment.module_id != module.id:
            errors.append(f"Module '{module.code or module.id}' does not match the assessment")
        if module and assessment and assessment.course_id and module.course_id != assessment.course_id:
            errors.append(f"Module '{module.code or module.id}' does not belong to the assessment course")
        if module and subject and subject.module_id != module.id:
            errors.append(f"Subject '{subject.code or subject.id}' does not belong to the module")

        if student and assessment:
            enrollment_query = db.session.query(Enrollment.id).filter(
                Enrollment.student_id == student.id,
                Enrollment.course_id == assessment.course_id,
                Enrollment.deleted_at.is_(None),
            )
            if module:
                enrollment_query = enrollment_query.filter(Enrollment.module_id == module.id)
            if not enrollment_query.first():
                errors.append("Student is not enrolled in the assessment module")

        # Compute grade
        grade = None
        is_passed = None
        if marks is not None and assessment:
            grade = _grade(marks, assessment.total_marks)
            is_passed = marks >= (assessment.pass_marks or assessment.total_marks * 0.5)

        # Flag learners the commit will pull onto the subject, so attaching them
        # is visible up front rather than a silent side effect of the upload.
        subject_link_missing = bool(
            student and subject and not student_subject_link_exists(student.id, subject.id)
        )

        # Flag rows that would overwrite a mark, so the uploader sees what is at
        # stake before committing rather than after.
        existing_score = None
        if student and assessment:
            existing_score = (
                db.session.query(Score)
                .filter_by(student_id=student.id, assessment_id=assessment.id)
                .filter(Score.deleted_at.is_(None))
                .first()
            )

        rows.append({
            "subject_link_missing": subject_link_missing,
            "has_existing_score": existing_score is not None,
            "existing_marks": existing_score.marks_obtained if existing_score else None,
            "existing_grade": existing_score.grade if existing_score else None,
            "row": i,
            "student_id": student_key,
            "student_code": student.code if student else None,
            "student_name": student.user.name if student and student.user else None,
            "registration_number": student.registration_number if student else student_key,
            "assessment_id": str(assessment.id) if assessment else assessment_id_str,
            "assessment_code": assessment.code if assessment else None,
            "assessment_name": assessment.name if assessment else None,
            "module_id": str(module.id) if module else (module_str or None),
            "module_code": module.code if module else None,
            "module_name": module.name if module else None,
            "subject_id": str(subject.id) if subject else (subject_str or None),
            "subject_code": subject.code if subject else None,
            "subject_name": subject.name if subject else None,
            "marks_obtained": marks,
            "total_marks": assessment.total_marks if assessment else None,
            "cat_formula": formula_details,
            "grade": grade,
            "is_passed": is_passed,
            "term": term,
            "feedback": feedback,
            "errors": errors,
            "valid": len(errors) == 0,
        })

    valid_count = sum(1 for r in rows if r["valid"])
    existing_count = sum(1 for r in rows if r["valid"] and r["has_existing_score"])
    link_count = len({
        r["student_id"] for r in rows if r["valid"] and r["subject_link_missing"]
    })
    return {
        "total": len(rows),
        "valid": valid_count,
        "invalid": len(rows) - valid_count,
        "existing": existing_count,
        # Distinct learners the commit will attach to the subject.
        "subject_links_missing": link_count,
        "rows": rows,
        "subject": {
            "id": str(batch_subject.id),
            "code": batch_subject.code,
            "name": batch_subject.name,
        } if batch_subject else None,
    }, 200


# ── Commit ────────────────────────────────────────────────────────────────────

@bp.post("/commit")
def commit_bulk():
    user, error, status = require_permission("scores.create")
    if error:
        return error, status

    batch_subject, subject_error, subject_status = _selected_subject(user)
    if subject_error:
        return subject_error, subject_status

    evidence_files = usable_score_evidence_files(request.files.getlist("exam_copies"))
    if not (request.content_type and request.content_type.startswith("multipart/form-data")):
        return {"error": "Use multipart/form-data"}, 400

    try:
        payload = {"rows": json.loads(request.form.get("rows", "[]"))}
    except json.JSONDecodeError:
        return {"error": "'rows' must be valid JSON"}, 400

    rows = payload.get("rows", [])
    if not rows:
        return {"error": "No rows provided"}, 400

    mode = resolve_conflict_mode(request.form.get("on_conflict") or request.args.get("on_conflict"))

    batch_id = uuid.uuid4().hex
    inserted = 0
    updated = 0
    skipped = 0
    subjects_linked = 0
    errors = []
    conflicts: list[dict] = []
    assessment_ids = set()
    subject_ids = set()
    # One lookup per module, not per row — a large upload is otherwise a query
    # per learner for a fact that cannot change during the batch.
    sole_subject_cache: dict[uuid.UUID, uuid.UUID | None] = {}
    # Built once: {normalised name -> canonical name}, so a row's label can be
    # matched without a query per row.
    term_lookup = {
        (name or "").strip().lower(): name for name in _known_term_names()
    }

    for row in rows:
        # Never trust preview-only flags or calculated values from the client.
        # Re-resolve and validate every row during the write transaction.
        if not row.get("valid"):
            skipped += 1
            continue

        student_key = row.get("student_id") or row.get("registration_number", "")
        student = _resolve_student(student_key)
        if not student:
            errors.append(f"Row {row.get('row')}: student '{student_key}' not found")
            skipped += 1
            continue

        try:
            assessment_uuid = uuid.UUID(row["assessment_id"])
        except (ValueError, TypeError, KeyError):
            skipped += 1
            continue

        assessment = db.session.get(Assessment, assessment_uuid)
        if not assessment:
            skipped += 1
            continue
        if assessment.assessment_scope != "formative":
            errors.append(f"Row {row.get('row')}: summative/external assessment evidence is not permitted")
            skipped += 1
            continue
        assessment_ids.add(assessment_uuid)

        raw_module = str(row.get("module_id") or row.get("module_code") or "").strip()
        module = _resolve_module(raw_module) if raw_module else assessment.module
        if raw_module and (not module or module.deleted_at):
            errors.append(f"Row {row.get('row')}: module '{raw_module}' not found")
            skipped += 1
            continue
        if module and assessment.module_id and assessment.module_id != module.id:
            errors.append(f"Row {row.get('row')}: module does not match the assessment")
            skipped += 1
            continue
        if module and assessment.course_id and module.course_id != assessment.course_id:
            errors.append(f"Row {row.get('row')}: module does not belong to the assessment course")
            skipped += 1
            continue

        try:
            marks = float(row["marks_obtained"])
        except (TypeError, ValueError):
            skipped += 1
            errors.append(f"Row {row.get('row')}: marks_obtained must be numeric")
            continue
        if marks < 0 or marks > assessment.total_marks:
            errors.append(f"marks_obtained must be 0–{assessment.total_marks}")
            skipped += 1
            continue
        grade = _grade(marks, assessment.total_marks)
        is_passed = marks >= (assessment.pass_marks or assessment.total_marks * 0.5)
        # Snap the uploaded label onto a real term. A label that matches no
        # term is not merely cosmetic: every report filters by term, so those
        # marks land in the database and appear in none of them. Matching is
        # case- and whitespace-insensitive, and an unrecognised label falls back
        # to the assessment's own term rather than being stored as typed.
        term = _normalise_term(
            row.get("term"),
            assessment,
            term_lookup,
            row.get("row"),
            errors,
        )
        feedback = row.get("feedback") or None

        # Resolve subject_id (may be a UUID string or a code from preview output).
        # Re-resolve rather than trusting the client, and fall back to the
        # subject the uploader selected for the whole batch.
        subject_id = None
        raw_subject = str(row.get("subject_id") or "").strip()
        if raw_subject:
            row_subject = _resolve_subject(raw_subject)
            if not row_subject:
                errors.append(f"Row {row.get('row')}: subject '{raw_subject}' not found")
                skipped += 1
                continue
            subject_id = row_subject.id
        elif batch_subject:
            subject_id = batch_subject.id
        if not subject_id and assessment.module_id:
            # Neither the row nor the batch named a subject. Every report that
            # groups marks by class keys on `Score.subject_id`, so leaving it
            # null makes an upload land in the database and vanish from the
            # reports. When the assessment's module owns exactly one subject
            # there is no ambiguity — attribute it.
            subject_id = _sole_subject_of_module(assessment.module_id, sole_subject_cache)

        if subject_id:
            subject_ids.add(subject_id)
            row_subject = db.session.get(Subject, subject_id)
            if module and row_subject and row_subject.module_id != module.id:
                errors.append(f"Row {row.get('row')}: subject does not belong to the module")
                skipped += 1
                continue
            if not module and row_subject:
                module = row_subject.module

        enrollment_query = db.session.query(Enrollment).filter(
            Enrollment.student_id == student.id,
            Enrollment.course_id == assessment.course_id,
            Enrollment.deleted_at.is_(None),
        )
        if module:
            enrollment_query = enrollment_query.filter(Enrollment.module_id == module.id)
        enrollment = enrollment_query.first()
        if not enrollment:
            errors.append(f"Row {row.get('row')}: student is not enrolled in the assessment module")
            skipped += 1
            continue
        trainer = db.session.query(Trainer).filter(Trainer.user_id == user.id).first()
        if trainer and not _is_admin(user):
            has_access = db.session.query(TrainerSubject).filter(
                TrainerSubject.trainer_id == trainer.id,
                TrainerSubject.subject_id == subject_id,
            ).first() if subject_id else None
            if not has_access:
                from ..models.trainer_course import TrainerCourse
                has_access = db.session.query(TrainerCourse).filter(
                    TrainerCourse.trainer_id == trainer.id,
                    TrainerCourse.course_id == assessment.course_id,
                ).first()
            if not has_access:
                errors.append(f"Row {row.get('row')}: you do not have access to this assessment")
                skipped += 1
                continue

        # The row is accepted for this subject, so the learner takes it. Attach
        # them if the roster was missing the link, otherwise the mark lands but
        # the learner stays invisible on every subject and dashboard roster.
        if subject_id and link_student_to_subject(student.id, subject_id):
            subjects_linked += 1

        existing = (
            db.session.query(Score)
            .filter_by(student_id=student.id, assessment_id=assessment_uuid)
            .first()
        )

        if existing:
            # An existing mark is never overwritten on the uploader's behalf.
            # Without an explicit decision the whole batch is rolled back below
            # and the conflicts are handed back for confirmation.
            if mode is None:
                conflicts.append({
                    "row": row.get("row"),
                    "registration_number": student.registration_number,
                    "name": student.user.name if student.user else None,
                    "assessment_name": assessment.name,
                    "current_marks": existing.marks_obtained,
                    "current_grade": existing.grade,
                    "new_marks": marks,
                    "new_grade": grade,
                    "message": "This learner already has a mark for this assessment",
                })
                continue
            if mode == "skip":
                skipped += 1
                continue
            existing.marks_obtained = marks
            existing.grade = grade
            existing.is_passed = is_passed
            existing.term = term
            existing.feedback = feedback
            if subject_id:
                existing.subject_id = subject_id
            updated += 1
        else:
            db.session.add(Score(
                student_id=student.id,
                assessment_id=assessment_uuid,
                subject_id=subject_id,
                enrollment_id=enrollment.id,
                marks_obtained=marks,
                grade=grade,
                is_passed=is_passed,
                term=term,
                feedback=feedback,
            ))
            inserted += 1

    # Overwrites need consent, so a batch carrying any is discarded entirely and
    # handed back for a decision — no partial write, no evidence files stored.
    if conflicts:
        db.session.rollback()
        return conflict_response(conflicts, len(rows), "marks")

    try:
        saved_evidence = save_score_evidence_files(
            evidence_files,
            uploaded_by=user.id,
            batch_id=batch_id,
            assessment_id=next(iter(assessment_ids)) if len(assessment_ids) == 1 else None,
            subject_id=next(iter(subject_ids)) if len(subject_ids) == 1 else None,
        )
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        remove_score_evidence_files(locals().get("saved_evidence", []))
        return {"error": "Marks could not be saved"}, 409

    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        # Learners this upload attached to the subject because the roster was
        # missing them.
        "subjects_linked": subjects_linked,
        "errors": errors,
        "on_conflict": mode,
        "batch_id": batch_id,
        "evidence_files": len(evidence_files),
        "subject": {
            "id": str(batch_subject.id),
            "code": batch_subject.code,
            "name": batch_subject.name,
        } if batch_subject else None,
    }, 200


# ── Template download ─────────────────────────────────────────────────────────

TEMPLATE_HEADER = [
    "student_id",
    "student_name",
    "marks_obtained",
    "assessment_code",
    "module_code",
    "subject_code",
    "term",
    "feedback",
]

CAT_FORMULA_TEMPLATE_HEADER = [
    "student_id",
    "student_name",
    "cat_1_marks",
    "cat_2_marks",
    "cat_3_marks",
    "assessment_code",
    "module_code",
    "subject_code",
    "term",
    "feedback",
]


def _known_term_names() -> list[str]:
    """Every term name, newest first — the only values the term column accepts."""
    return [
        row[0]
        for row in db.session.query(Term.name)
        .filter(Term.deleted_at.is_(None))
        .order_by(Term.start_date.desc())
        .all()
        if row[0]
    ]


def _template_term(assessment: Assessment | None) -> str:
    """
    The term to prefill.

    A blank term column is where the whole problem starts: the uploader types
    their own label, it matches no term, and every term-scoped report silently
    omits the marks. Prefer the assessment's own term, fall back to the current
    one, so the column is never empty for someone to fill in freehand.
    """
    if assessment and assessment.term and assessment.term.name:
        return assessment.term.name
    active = (
        db.session.query(Term)
        .filter(Term.is_active.is_(True), Term.deleted_at.is_(None))
        .first()
    )
    if active and active.name:
        return active.name
    names = _known_term_names()
    return names[0] if names else ""


def _restrict_term_column(worksheet, header: list[str], row_count: int) -> None:
    """
    Turn the term column into a dropdown of real terms.

    Prefilling stops the common case; this stops the rest, because a value
    typed over the prefill is exactly as invisible as a blank one was.
    """
    names = _known_term_names()
    if not names or "term" not in header:
        return
    column = get_column_letter(header.index("term") + 1)
    # Excel caps an inline list at 255 characters; beyond that the sheet still
    # opens but the dropdown is dropped, so only apply it when it fits.
    formula = ",".join(name.replace('"', "'") for name in names)
    if len(formula) > 250:
        return
    validation = DataValidation(
        type="list",
        formula1=f'"{formula}"',
        allow_blank=False,
        showDropDown=False,
    )
    validation.error = (
        "Pick a term from the list. A term typed by hand that does not match "
        "one of these hides the marks from every report."
    )
    validation.errorTitle = "Unknown term"
    validation.prompt = "Choose the term these marks belong to."
    validation.promptTitle = "Term"
    worksheet.add_data_validation(validation)
    validation.add(f"{column}2:{column}{max(row_count + 1, 200)}")


def _template_roster(user, assessment: Assessment, subject: Subject | None) -> list[Student]:
    """
    Learners who may legitimately receive marks for this assessment.

    Mirrors the rules `commit_bulk` enforces — enrolled in the assessment's
    course, and within the uploader's own subjects when they are a trainer — so
    a prefilled row never turns into a rejected one.

    A selected subject with nobody on its roster yet is the one case the class
    list cannot come from: it falls back to the learners enrolled in the
    assessment's course, which is how a subject taught to a course that was
    never split into subject rosters gets its first class list. Committing the
    marks attaches them, so the fallback only ever applies once.
    """
    query = (
        db.session.query(Student)
        .join(Enrollment, Enrollment.student_id == Student.id)
        .filter(
            Enrollment.deleted_at.is_(None),
            Student.deleted_at.is_(None),
        )
    )

    if assessment.course_id:
        query = query.filter(Enrollment.course_id == assessment.course_id)
    if assessment.module_id:
        query = query.filter(Enrollment.module_id == assessment.module_id)

    subject_ids: set[uuid.UUID] | None = None
    if subject:
        subject_ids = {subject.id}
    else:
        trainer = _uploader_trainer(user)
        if trainer:
            subject_ids = _trainer_subject_ids(trainer)
            if not subject_ids:
                return []

    if not assessment.course_id and subject_ids is None:
        # Nothing to scope by — refuse to dump every learner in the institution.
        return []

    scoped_query = query
    if subject_ids is not None:
        scoped_query = query.join(
            StudentSubject, StudentSubject.student_id == Student.id
        ).filter(StudentSubject.subject_id.in_(subject_ids))

    roster = scoped_query.distinct().order_by(Student.code.asc()).all()
    if roster or not subject or not assessment.course_id:
        return roster

    return query.distinct().order_by(Student.code.asc()).all()


def _template_subject_reference(user, assessment: Assessment | None) -> list[list[str]]:
    """Subject reference rows visible to the uploader, optionally scoped by course."""
    query = (
        db.session.query(Subject)
        .join(Module, Module.id == Subject.module_id)
        .join(Course, Course.id == Module.course_id)
        .join(Department, Department.id == Course.department_id)
        .filter(
            Subject.deleted_at.is_(None),
            Module.deleted_at.is_(None),
            Course.deleted_at.is_(None),
        )
    )
    if assessment and assessment.course_id:
        query = query.filter(Module.course_id == assessment.course_id)
    if user.institution_id:
        query = query.filter(Department.institution_id == user.institution_id)

    trainer = _uploader_trainer(user)
    if trainer:
        subject_ids = _trainer_subject_ids(trainer)
        if not subject_ids:
            return []
        query = query.filter(Subject.id.in_(subject_ids))

    rows = []
    for subject in query.order_by(Course.name, Module.name, Subject.name).all():
        module = subject.module
        course = module.course if module else None
        rows.append([
            module.code or str(module.id) if module else "",
            module.name if module else "",
            course.code or str(course.id) if course else "",
            course.name if course else "",
            subject.code or str(subject.id),
            subject.name,
        ])
    return rows


def _template_module_reference(user, assessment: Assessment | None) -> list[list[str]]:
    """Current modules visible to the uploader, optionally scoped by assessment."""
    query = (
        db.session.query(Module)
        .join(Course, Course.id == Module.course_id)
        .join(Department, Department.id == Course.department_id)
        .filter(
            Module.deleted_at.is_(None),
            Course.deleted_at.is_(None),
        )
    )
    if assessment and assessment.course_id:
        query = query.filter(Module.course_id == assessment.course_id)
    if user.institution_id:
        query = query.filter(Department.institution_id == user.institution_id)

    trainer = _uploader_trainer(user)
    if trainer:
        subject_ids = _trainer_subject_ids(trainer)
        if not subject_ids:
            return []
        query = query.join(Subject, Subject.module_id == Module.id).filter(
            Subject.id.in_(subject_ids),
            Subject.deleted_at.is_(None),
        )

    rows = []
    seen_module_ids = set()
    for module in query.order_by(Course.name, Module.name).all():
        if module.id in seen_module_ids:
            continue
        seen_module_ids.add(module.id)
        rows.append([
            module.code or str(module.id),
            module.name,
            module.course.code or str(module.course.id) if module.course else "",
            module.course.name if module.course else "",
        ])
    return rows


def _format_template_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_number, cells in enumerate(worksheet.columns, start=1):
        width = min(max(len(str(cell.value or "")) for cell in cells) + 2, 42)
        worksheet.column_dimensions[get_column_letter(column_number)].width = width


@bp.get("/template")
def download_template():
    """
    XLSX template for a marks upload.

    With `assessment_id` (and optionally `subject_code`) it comes back prefilled
    with the class list — one row per learner, marks left blank. The second sheet
    lists current modules available to the uploader.
    """
    user, error, status = require_permission("scores.create")
    if error:
        return error, status

    batch_subject, subject_error, subject_status = _selected_subject(user)
    if subject_error:
        return subject_error, subject_status

    cat_formula_key = (request.args.get("cat_formula") or "").strip().lower()
    cat_formula = _resolve_cat_formula(cat_formula_key)
    if cat_formula_key and not cat_formula:
        return {"error": "Unknown CAT formula scenario"}, 400

    assessment_ref = (
        request.args.get("assessment_id")
        or request.args.get("assessment_code")
        or ""
    ).strip()
    assessment = _resolve_assessment(assessment_ref) if assessment_ref else None
    if assessment_ref and not assessment:
        return {"error": f"Assessment '{assessment_ref}' not found"}, 404

    workbook = Workbook()
    marks_sheet = workbook.active
    marks_sheet.title = "Marks Upload"
    template_header = CAT_FORMULA_TEMPLATE_HEADER if cat_formula else TEMPLATE_HEADER
    marks_sheet.append(template_header)
    term = _template_term(assessment)

    learner_rows = 0
    if assessment:
        subject_code = batch_subject.code if batch_subject else ""
        module = batch_subject.module if batch_subject else assessment.module
        module_code = module.code if module else ""
        for student in _template_roster(user, assessment, batch_subject):
            common = [
                student.code or student.registration_number,
                student.user.name if student.user else "",
            ]
            if cat_formula:
                marks_sheet.append([
                    *common,
                    "",
                    "",
                    "",
                    # Use the human-readable assessment code in exported files.
                    # Legacy records without a code fall back to their UUID.
                    assessment.code or str(assessment.id),
                    module_code,
                    subject_code,
                    term,
                    "",
                ])
            else:
                marks_sheet.append([
                    *common,
                    "",  # marks_obtained — the one column to fill in
                    assessment.code or str(assessment.id),
                    module_code,
                    subject_code,
                    term,
                    "",
                ])
            learner_rows += 1
        filename = f"marks_template_{assessment.code or 'assessment'}.xlsx"
    else:
        # A generic template must not contain made-up identifiers that look
        # uploadable, but the term is a real value and the one most often typed
        # wrong, so it is filled in even here.
        blank_row = ["" for _ in template_header]
        blank_row[template_header.index("term")] = term
        marks_sheet.append(blank_row)
        filename = "marks_upload_template.xlsx"

    _restrict_term_column(marks_sheet, template_header, learner_rows)
    _format_template_sheet(marks_sheet)

    term_sheet = workbook.create_sheet("Terms")
    term_sheet.append(["term", "status"])
    active_name = _template_term(assessment)
    for name in _known_term_names():
        term_sheet.append([name, "current" if name == active_name else ""])
    if not _known_term_names():
        term_sheet.append(["(no terms defined — create one under Academic Terms)", ""])
    _format_template_sheet(term_sheet)

    module_sheet = workbook.create_sheet("Modules")
    module_sheet.append([
        "module_code",
        "module_name",
        "course_code",
        "course_name",
    ])
    for reference_row in _template_module_reference(user, assessment):
        module_sheet.append(reference_row)
    _format_template_sheet(module_sheet)

    reference_sheet = workbook.create_sheet("Subject Codes")
    reference_sheet.append([
        "module_code",
        "module_name",
        "course_code",
        "course_name",
        "subject_code",
        "subject_name",
    ])
    for reference_row in _template_subject_reference(user, assessment):
        reference_sheet.append(reference_row)
    _format_template_sheet(reference_sheet)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
    response.headers["X-Template-Rows"] = str(learner_rows)
    response.headers["X-Template-Prefilled"] = "1" if assessment else "0"
    if cat_formula:
        response.headers["X-CAT-Formula"] = cat_formula_key
    return response


@bp.get("/evidence/files/<path:filename>")
def serve_evidence_file(filename: str):
    user, error, status = get_current_user()
    if error:
        return error, status
    evidence = db.session.query(ScoreEvidence).filter(
        ScoreEvidence.file_url == f"/scores/evidence/files/{filename}"
    ).first()
    if not evidence or not can_access_score_evidence(user, evidence):
        return {"error": "File not found"}, 404
    return send_from_directory(EVIDENCE_UPLOAD_FOLDER, filename)
