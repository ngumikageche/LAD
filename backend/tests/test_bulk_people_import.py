import io

import pytest
from openpyxl import Workbook, load_workbook
from werkzeug.datastructures import FileStorage

from app.services.bulk_people_import import (
    build_template,
    first_value,
    normalize_lookup,
    read_people_upload,
)


def workbook_upload(headers, row, sheet_name="Students Template"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    worksheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return FileStorage(
        stream=output,
        filename="import.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def test_import_recognizes_useful_columns_from_external_student_workbook():
    upload = workbook_upload(
        ["Reg No", "Name", "Email", "Mobile", "Course Code", "Date Of Admission(dd/MM/yyyy)"],
        ["TVET-1", "Amina", "amina@example.edu", "0712345678", "CRS001", "15/01/2026"],
    )

    rows = read_people_upload(upload, preferred_sheet="Students Template")

    assert first_value(rows[0], "Registration Number", "Reg No") == "TVET-1"
    assert first_value(rows[0], "Course Code") == "CRS001"


def test_csv_import_ignores_blank_rows_and_normalizes_headers():
    upload = FileStorage(
        stream=io.BytesIO(b"Name,Department\nTrainer One,Electrical\n,\n"),
        filename="trainers.csv",
        content_type="text/csv",
    )

    rows = read_people_upload(upload)

    assert rows == [{"name": "Trainer One", "department": "Electrical"}]


def test_import_rejects_unsupported_file_type():
    upload = FileStorage(stream=io.BytesIO(b"data"), filename="people.pdf")

    with pytest.raises(ValueError, match="CSV or XLSX"):
        read_people_upload(upload)


def test_generated_template_contains_only_our_selected_columns():
    output = build_template(
        ["Registration Number", "Name", "Email", "Course Code"],
        "Learners",
        reference_sheets={
            "Course Codes": (
                ["Course Code", "Course Name", "Institution Name"],
                [["CRS001", "Electrical Level 6", "LAD College"]],
            ),
        },
    )

    workbook = load_workbook(output, read_only=True)
    headers = [cell.value for cell in next(workbook["Learners"].iter_rows())]

    assert headers == ["Registration Number", "Name", "Email", "Course Code"]
    assert [cell.value for cell in next(workbook["Course Codes"].iter_rows())] == [
        "Course Code",
        "Course Name",
        "Institution Name",
    ]


def test_lookup_normalization_tolerates_extra_institutional_spacing():
    assert normalize_lookup("  DIPLOMA   IN  ICT (KNEC) ") == normalize_lookup("Diploma in ICT (KNEC)")
