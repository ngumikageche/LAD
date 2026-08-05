import io
import json
import uuid

from openpyxl import load_workbook
from werkzeug.security import generate_password_hash


def _login(client, email: str) -> dict:
    response = client.post("/auth/login", json={"email": email, "password": "S3cret!"})
    token = response.get_json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _seed(app):
    """
    One course with two subjects. The trainer is assigned to SUB-A only.
    Two learners are enrolled in the course, but only the first takes SUB-A —
    so subject scoping has something to filter out.
    """
    from datetime import datetime

    from app.extensions import db
    from app.models.assessment import Assessment
    from app.models.course import Course
    from app.models.department import Department
    from app.models.enrollment import Enrollment
    from app.models.institution import Institution
    from app.models.module import Module
    from app.models.role_permission import RolePermission
    from app.models.student import Student
    from app.models.student_subject import StudentSubject
    from app.models.subject import Subject
    from app.models.term import Term
    from app.models.trainer import Trainer
    from app.models.trainer_subject import TrainerSubject
    from app.models.user import User

    with app.app_context():
        admin_role = RolePermission(role_name="Admin", permissions={"*": True})
        trainer_role = RolePermission(
            role_name="Trainer",
            permissions={"scores.create": True, "scores.read": True},
        )
        institution = Institution(name="LAD College", type="College", location="Nairobi")
        db.session.add_all([admin_role, trainer_role, institution])
        db.session.flush()

        department = Department(institution_id=institution.id, name="Engineering")
        db.session.add(department)
        db.session.flush()

        course = Course(department_id=department.id, name="Electrical", cbet_level="Level 5")
        db.session.add(course)
        db.session.flush()

        module = Module(course_id=course.id, name="Circuits")
        db.session.add(module)
        db.session.flush()

        assigned_subject = Subject(module_id=module.id, name="Circuit Analysis")
        unassigned_subject = Subject(module_id=module.id, name="Thermodynamics")
        db.session.add_all([assigned_subject, unassigned_subject])
        db.session.flush()

        term = Term(
            name="Term 1 2026",
            start_date=datetime(2026, 1, 1),
            end_date=datetime(2026, 4, 1),
            is_active=True,
        )
        db.session.add(term)
        db.session.flush()

        assessment = Assessment(
            course_id=course.id,
            module_id=module.id,
            term_id=term.id,
            name="CAT 1",
            assessment_type="quiz",
            assessment_scope="formative",
            total_marks=100,
            pass_marks=50,
        )
        db.session.add(assessment)
        db.session.flush()

        admin_user = User(
            name="Admin",
            email="admin@example.com",
            password_hash=generate_password_hash("S3cret!"),
            role_id=admin_role.id,
            institution_id=institution.id,
        )
        trainer_user = User(
            name="Trainer One",
            email="trainer@example.com",
            password_hash=generate_password_hash("S3cret!"),
            role_id=trainer_role.id,
            institution_id=institution.id,
        )
        student_user = User(
            name="Learner One",
            email="learner@example.com",
            password_hash=generate_password_hash("S3cret!"),
            role_id=admin_role.id,
            institution_id=institution.id,
        )
        other_student_user = User(
            name="Learner Two",
            email="learner2@example.com",
            password_hash=generate_password_hash("S3cret!"),
            role_id=admin_role.id,
            institution_id=institution.id,
        )
        db.session.add_all([admin_user, trainer_user, student_user, other_student_user])
        db.session.flush()

        trainer = Trainer(user_id=trainer_user.id, department_id=department.id)
        student = Student(
            user_id=student_user.id,
            registration_number="REG-001",
            course_id=course.id,
            enrollment_year=2026,
        )
        other_student = Student(
            user_id=other_student_user.id,
            registration_number="REG-002",
            course_id=course.id,
            enrollment_year=2026,
        )
        db.session.add_all([trainer, student, other_student])
        db.session.flush()

        db.session.add_all([
            TrainerSubject(trainer_id=trainer.id, subject_id=assigned_subject.id),
            StudentSubject(student_id=student.id, subject_id=assigned_subject.id),
            StudentSubject(student_id=other_student.id, subject_id=unassigned_subject.id),
            Enrollment(student_id=student.id, course_id=course.id, module_id=module.id),
            Enrollment(student_id=other_student.id, course_id=course.id, module_id=module.id),
        ])
        db.session.commit()

        return {
            "assessment_id": str(assessment.id),
            "assessment_code": assessment.code,
            "course_code": course.code,
            "course_name": course.name,
            "module_code": module.code,
            "module_name": module.name,
            "student_code": student.code,
            "student_id": str(student.id),
            "other_student_id": str(other_student.id),
            "assigned_subject_code": assigned_subject.code,
            "assigned_subject_id": str(assigned_subject.id),
            "unassigned_subject_code": unassigned_subject.code,
            "term_name": term.name,
            "other_student_code": other_student.code,
        }


def _csv(
    student_code: str,
    assessment_code: str,
    subject_code: str = "",
    module_code: str = "",
) -> tuple:
    body = "student_id,marks_obtained,assessment_id,module_code,subject_id\n"
    body += f"{student_code},75,{assessment_code},{module_code},{subject_code}\n"
    return (io.BytesIO(body.encode("utf-8")), "marks.csv")


def test_admin_lists_every_subject_by_code(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    response = client.get("/scores/bulk-marks/subjects", headers=headers)

    assert response.status_code == 200
    body = response.get_json()
    assert body["scope"] == "all"
    codes = {subject["code"] for subject in body["subjects"]}
    assert codes == {ids["assigned_subject_code"], ids["unassigned_subject_code"]}


def test_trainer_only_lists_subjects_they_are_assigned(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    body = client.get("/scores/bulk-marks/subjects", headers=headers).get_json()

    assert body["scope"] == "trainer"
    assert [subject["code"] for subject in body["subjects"]] == [ids["assigned_subject_code"]]


def test_trainer_can_create_formative_assessment_for_assigned_subject(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    response = client.post(
        "/scores/bulk-marks/assessments",
        headers=headers,
        json={
            "name": "CAT 2",
            "assessment_type": "test",
            "subject_code": ids["assigned_subject_code"],
            "total_marks": 50,
            "pass_marks": 25,
        },
    )

    assert response.status_code == 201
    body = response.get_json()
    assert body["code"].startswith("ASM")
    assert body["assessment_scope"] == "formative"
    assert body["subject_code"] == ids["assigned_subject_code"]
    assert body["total_marks"] == 50


def test_trainer_cannot_create_assessment_for_unassigned_subject(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    response = client.post(
        "/scores/bulk-marks/assessments",
        headers=headers,
        json={
            "name": "Unauthorized CAT",
            "subject_code": ids["unassigned_subject_code"],
            "total_marks": 100,
            "pass_marks": 50,
        },
    )

    assert response.status_code == 403


def test_selected_subject_code_fills_rows_that_omit_one(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    response = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["student_code"], ids["assessment_code"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    body = response.get_json()
    assert body["subject"]["code"] == ids["assigned_subject_code"]
    row = body["rows"][0]
    assert row["valid"] is True
    assert row["subject_id"] == ids["assigned_subject_id"]
    assert row["subject_code"] == ids["assigned_subject_code"]


def test_row_level_subject_code_overrides_the_batch_subject(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    body = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(
                ids["student_code"], ids["assessment_code"], ids["unassigned_subject_code"]
            ),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()

    assert body["rows"][0]["subject_code"] == ids["unassigned_subject_code"]


def test_unknown_subject_code_is_rejected(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    response = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["student_code"], ids["assessment_code"]),
            "subject_code": "SUB-NOPE",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 404


def test_unknown_row_module_code_makes_the_marks_row_invalid(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    response = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(
                ids["student_code"],
                ids["assessment_code"],
                ids["assigned_subject_code"],
                "MOD-NOPE",
            ),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    row = response.get_json()["rows"][0]
    assert row["valid"] is False
    assert "Module 'MOD-NOPE' not found" in row["errors"]


def test_trainer_cannot_select_a_subject_they_do_not_teach(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    response = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["student_code"], ids["assessment_code"]),
            "subject_code": ids["unassigned_subject_code"],
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 403


def _load_template(response):
    return load_workbook(io.BytesIO(response.get_data()), data_only=True)


def _sheet_rows(workbook, sheet_name: str):
    values = workbook[sheet_name].iter_rows(values_only=True)
    headers = next(values)
    return [
        {header: "" if value is None else str(value) for header, value in zip(headers, row)}
        for row in values
    ]


def test_template_prefills_the_class_list_for_an_assessment(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    response = client.get(
        f"/scores/bulk-marks/template?assessment_id={ids['assessment_code']}",
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["X-Template-Prefilled"] == "1"
    assert response.headers["X-Template-Rows"] == "2"
    assert ids["assessment_code"] in response.headers["Content-Disposition"]

    workbook = _load_template(response)
    assert workbook.sheetnames == ["Marks Upload", "Modules", "Subject Codes"]
    rows = _sheet_rows(workbook, "Marks Upload")
    assert {row["student_id"] for row in rows} == {
        ids["student_code"],
        ids["other_student_code"],
    }
    assert {row["student_name"] for row in rows} == {"Learner One", "Learner Two"}
    # Marks are the only thing left to fill in.
    assert all(row["marks_obtained"] == "" for row in rows)
    assert all(row["assessment_code"] == ids["assessment_code"] for row in rows)
    assert all(row["module_code"] == ids["module_code"] for row in rows)
    assert all(row["term"] == ids["term_name"] for row in rows)

    module_rows = _sheet_rows(workbook, "Modules")
    assert {
        "module_code": ids["module_code"],
        "module_name": ids["module_name"],
        "course_code": ids["course_code"],
        "course_name": ids["course_name"],
    } in module_rows

    reference_rows = _sheet_rows(workbook, "Subject Codes")
    assert {
        "module_code": ids["module_code"],
        "module_name": ids["module_name"],
        "course_code": ids["course_code"],
        "course_name": ids["course_name"],
        "subject_code": ids["assigned_subject_code"],
        "subject_name": "Circuit Analysis",
    } in reference_rows


def test_template_narrows_the_class_list_to_the_selected_subject(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    response = client.get(
        f"/scores/bulk-marks/template?assessment_id={ids['assessment_code']}"
        f"&subject_code={ids['assigned_subject_code']}",
        headers=headers,
    )

    rows = _sheet_rows(_load_template(response), "Marks Upload")
    assert [row["student_id"] for row in rows] == [ids["student_code"]]
    assert rows[0]["subject_code"] == ids["assigned_subject_code"]


def test_trainer_template_only_lists_learners_in_their_own_subjects(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    response = client.get(
        f"/scores/bulk-marks/template?assessment_id={ids['assessment_code']}",
        headers=headers,
    )

    # Learner Two takes only the subject this trainer does not teach.
    rows = _sheet_rows(_load_template(response), "Marks Upload")
    assert [row["student_id"] for row in rows] == [ids["student_code"]]


def test_template_without_an_assessment_does_not_export_a_fake_id(client, app):
    _seed(app)
    headers = _login(client, "admin@example.com")

    response = client.get("/scores/bulk-marks/template", headers=headers)

    assert response.status_code == 200
    assert response.headers["X-Template-Prefilled"] == "0"
    rows = _sheet_rows(_load_template(response), "Marks Upload")
    assert len(rows) == 1
    assert rows[0]["student_id"] == ""
    assert rows[0]["assessment_code"] == ""


def test_template_requires_authentication(client, app):
    _seed(app)

    # It now carries real learner names, so it must not be publicly downloadable.
    assert client.get("/scores/bulk-marks/template").status_code == 401


def test_prefilled_template_round_trips_through_preview(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    downloaded = client.get(
        f"/scores/bulk-marks/template?assessment_id={ids['assessment_code']}"
        f"&subject_code={ids['assigned_subject_code']}",
        headers=headers,
    )

    # Fill in the one blank column, exactly as a trainer would.
    workbook = _load_template(downloaded)
    worksheet = workbook["Marks Upload"]
    headers_row = [cell.value for cell in worksheet[1]]
    marks_column = headers_row.index("marks_obtained") + 1
    worksheet.cell(row=2, column=marks_column, value=64)
    filled = io.BytesIO()
    workbook.save(filled)
    filled.seek(0)

    body = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={"file": (filled, "marks.xlsx")},
        content_type="multipart/form-data",
    ).get_json()

    assert body["valid"] == 1
    row = body["rows"][0]
    assert row["marks_obtained"] == 64
    assert row["module_code"] == ids["module_code"]
    assert row["subject_code"] == ids["assigned_subject_code"]
    assert row["errors"] == []


def test_cat_formula_scenario_1_uses_30_40_30_totals(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")
    body = (
        "student_id,cat_1_marks,cat_2_marks,cat_3_marks,assessment_code,module_code,subject_code\n"
        f"{ids['student_code']},20,35,25,{ids['assessment_code']},{ids['module_code']},{ids['assigned_subject_code']}\n"
    )

    response = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": (io.BytesIO(body.encode("utf-8")), "marks.csv"),
            "cat_formula": "scenario_1",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200, response.get_json()
    row = response.get_json()["rows"][0]
    assert row["valid"] is True
    assert row["marks_obtained"] == 80
    assert row["cat_formula"]["component_total_marks"] == [30, 40, 30]
    assert row["cat_formula"]["final_percentage"] == 80


def test_cat_formula_scenario_3_uses_80_60_100_totals(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")
    body = (
        "student_id,cat_1_marks,cat_2_marks,cat_3_marks,assessment_code,module_code,subject_code\n"
        f"{ids['student_code']},40,30,50,{ids['assessment_code']},{ids['module_code']},{ids['assigned_subject_code']}\n"
    )

    response = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": (io.BytesIO(body.encode("utf-8")), "marks.csv"),
            "cat_formula": "scenario_3",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200, response.get_json()
    row = response.get_json()["rows"][0]
    assert row["valid"] is True
    assert row["marks_obtained"] == 50
    assert row["cat_formula"]["component_total_marks"] == [80, 60, 100]
    assert row["cat_formula"]["final_percentage"] == 50


def test_cat_formula_rejects_component_marks_over_their_maximum(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")
    body = (
        "student_id,cat_1_marks,cat_2_marks,cat_3_marks,assessment_code,module_code,subject_code\n"
        f"{ids['student_code']},31,35,25,{ids['assessment_code']},{ids['module_code']},{ids['assigned_subject_code']}\n"
    )

    response = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": (io.BytesIO(body.encode("utf-8")), "marks.csv"),
            "cat_formula": "scenario_1",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200, response.get_json()
    row = response.get_json()["rows"][0]
    assert row["valid"] is False
    assert "cat_1_marks must be 0–30" in row["errors"]


def test_commit_stores_the_selected_subject_without_requiring_exam_copies(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    preview = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["student_code"], ids["assessment_code"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()

    response = client.post(
        "/scores/bulk-marks/commit",
        headers=headers,
        data={
            "rows": json.dumps(preview["rows"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    body = response.get_json()
    assert body["inserted"] == 1
    assert body["evidence_files"] == 0
    assert body["subject"]["code"] == ids["assigned_subject_code"]

    from app.extensions import db
    from app.models.score import Score

    with app.app_context():
        score = db.session.query(Score).one()
        assert str(score.subject_id) == ids["assigned_subject_id"]
        assert score.marks_obtained == 75
        assert score.grade == "B"


def test_individual_scores_reuse_one_assessment_record(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    response = client.post(
        "/api/v1/trainer/scores",
        headers=headers,
        data={
            "student_id": ids["student_id"],
            "subject_id": ids["assigned_subject_id"],
            "assessment_id": ids["assessment_id"],
            "term": ids["term_name"],
            "score": "82",
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 201
    body = response.get_json()
    assert body["assessment_id"] == ids["assessment_id"]
    assert body["assessment"]["code"] == ids["assessment_code"]
    assert body["grade"] == "A"

    duplicate = client.post(
        "/api/v1/trainer/scores",
        headers=headers,
        data={
            "student_id": ids["student_id"],
            "subject_id": ids["assigned_subject_id"],
            "assessment_id": ids["assessment_id"],
            "term": ids["term_name"],
            "score": "75",
        },
        content_type="multipart/form-data",
    )

    assert duplicate.status_code == 409
    assert "already has a score" in duplicate.get_json()["error"]


def test_preview_flags_learners_missing_from_the_subject_roster(client, app):
    """Learner Two is enrolled in the course but not on the assessed subject."""
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    body = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["other_student_code"], ids["assessment_code"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()

    assert body["valid"] == 1
    assert body["subject_links_missing"] == 1
    assert body["rows"][0]["subject_link_missing"] is True


def test_preview_does_not_flag_learners_already_on_the_subject(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    body = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["student_code"], ids["assessment_code"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()

    assert body["subject_links_missing"] == 0
    assert body["rows"][0]["subject_link_missing"] is False


def test_commit_attaches_marked_learners_who_were_missing_the_subject(client, app):
    """
    The whole point: a learner who is marked for a subject must end up on its
    roster, or they stay invisible in every trainer count built from it.
    """
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    preview = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["other_student_code"], ids["assessment_code"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()

    response = client.post(
        "/scores/bulk-marks/commit",
        headers=headers,
        data={
            "rows": json.dumps(preview["rows"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    body = response.get_json()
    assert body["inserted"] == 1
    assert body["subjects_linked"] == 1

    from app.extensions import db
    from app.models.student_subject import StudentSubject

    with app.app_context():
        link = (
            db.session.query(StudentSubject)
            .filter(
                StudentSubject.subject_id == uuid.UUID(ids["assigned_subject_id"]),
                StudentSubject.student_id == uuid.UUID(ids["other_student_id"]),
            )
            .one_or_none()
        )
        assert link is not None


def test_commit_does_not_duplicate_an_existing_subject_link(client, app):
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    preview = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["student_code"], ids["assessment_code"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()

    body = client.post(
        "/scores/bulk-marks/commit",
        headers=headers,
        data={
            "rows": json.dumps(preview["rows"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()

    assert body["subjects_linked"] == 0

    from app.extensions import db
    from app.models.student_subject import StudentSubject

    with app.app_context():
        links = (
            db.session.query(StudentSubject)
            .filter(
                StudentSubject.subject_id == uuid.UUID(ids["assigned_subject_id"]),
                StudentSubject.student_id == uuid.UUID(ids["student_id"]),
            )
            .count()
        )
        assert links == 1


def test_marked_learners_start_counting_on_the_trainer_dashboard(client, app):
    """The dashboard total is built from the subject roster, so it must move."""
    ids = _seed(app)
    headers = _login(client, "trainer@example.com")

    before = client.get("/api/v1/trainer/dashboard", headers=headers).get_json()
    assert before["total_students"] == 1

    preview = client.post(
        "/scores/bulk-marks/preview",
        headers=headers,
        data={
            "file": _csv(ids["other_student_code"], ids["assessment_code"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    ).get_json()
    client.post(
        "/scores/bulk-marks/commit",
        headers=headers,
        data={
            "rows": json.dumps(preview["rows"]),
            "subject_code": ids["assigned_subject_code"],
        },
        content_type="multipart/form-data",
    )

    after = client.get("/api/v1/trainer/dashboard", headers=headers).get_json()
    assert after["total_students"] == 2


def test_template_falls_back_to_the_course_when_the_subject_has_no_roster(client, app):
    """
    A subject nobody is attached to yet cannot produce a class list, and an
    empty template would leave the trainer with no way to bootstrap one.
    """
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    from app.extensions import db
    from app.models.student_subject import StudentSubject

    with app.app_context():
        db.session.query(StudentSubject).filter(
            StudentSubject.subject_id == uuid.UUID(ids["assigned_subject_id"])
        ).delete(synchronize_session=False)
        db.session.commit()

    response = client.get(
        f"/scores/bulk-marks/template?assessment_id={ids['assessment_code']}"
        f"&subject_code={ids['assigned_subject_code']}",
        headers=headers,
    )

    rows = _sheet_rows(_load_template(response), "Marks Upload")
    assert {row["student_id"] for row in rows} == {
        ids["student_code"],
        ids["other_student_code"],
    }
    assert all(row["subject_code"] == ids["assigned_subject_code"] for row in rows)
