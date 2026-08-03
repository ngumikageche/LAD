import io

from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash


def _login(client, email: str) -> dict:
    response = client.post("/auth/login", json={"email": email, "password": "S3cret!"})
    return {"Authorization": f"Bearer {response.get_json()['access_token']}"}


def _seed(app):
    from app.extensions import db
    from app.models.course import Course
    from app.models.department import Department
    from app.models.institution import Institution
    from app.models.module import Module
    from app.models.role_permission import RolePermission
    from app.models.subject import Subject
    from app.models.user import User

    with app.app_context():
        admin_role = RolePermission(role_name="Admin", permissions={"data.import": True})
        student_role = RolePermission(role_name="Student", permissions={})
        institution = Institution(name="LAD College", type="College", location="Nairobi")
        db.session.add_all([admin_role, student_role, institution])
        db.session.flush()

        department = Department(institution_id=institution.id, name="Engineering")
        db.session.add(department)
        db.session.flush()
        course = Course(
            department_id=department.id,
            name="Electrical",
            cbet_level="Level 5",
        )
        db.session.add(course)
        db.session.flush()
        selected_module = Module(course_id=course.id, name="Circuits")
        other_module = Module(course_id=course.id, name="Machines")
        db.session.add_all([selected_module, other_module])
        db.session.flush()
        selected_subject = Subject(module_id=selected_module.id, name="Circuit Analysis")
        other_subject = Subject(module_id=other_module.id, name="Electrical Machines")
        db.session.add_all([selected_subject, other_subject])

        admin = User(
            name="Admin",
            email="admin@example.com",
            password_hash=generate_password_hash("S3cret!"),
            role_id=admin_role.id,
            institution_id=institution.id,
        )
        db.session.add(admin)
        db.session.commit()
        return {
            "course_code": course.code,
            "module_code": selected_module.code,
            "module_id": selected_module.id,
            "subject_id": selected_subject.id,
            "other_subject_id": other_subject.id,
        }


def _student_workbook(course_code: str, module_code: str) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Students Template"
    worksheet.append([
        "Registration Number",
        "Name",
        "Email",
        "Mobile",
        "Course Code",
        "Module Code",
        "Admission Date",
    ])
    worksheet.append([
        "TVET/2026/100",
        "Amina Learner",
        "amina@example.com",
        "0712345678",
        course_code,
        module_code,
        "15/01/2026",
    ])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_student_template_and_import_use_current_module_codes(client, app):
    ids = _seed(app)
    headers = _login(client, "admin@example.com")

    template_response = client.get("/students/import-template", headers=headers)

    assert template_response.status_code == 200
    workbook = load_workbook(io.BytesIO(template_response.get_data()), data_only=True)
    assert workbook.sheetnames == ["Students Template", "Modules", "Course Codes"]
    template_headers = [cell.value for cell in workbook["Students Template"][1]]
    assert "Module Code" in template_headers
    module_rows = list(workbook["Modules"].iter_rows(values_only=True))
    assert module_rows[0] == ("Module Code", "Module Name", "Course Code", "Course Name")
    assert any(row[0] == ids["module_code"] for row in module_rows[1:])

    upload_response = client.post(
        "/students/bulk-upload",
        headers=headers,
        data={
            "file": (
                _student_workbook(ids["course_code"], ids["module_code"]),
                "students.xlsx",
            ),
        },
        content_type="multipart/form-data",
    )

    assert upload_response.status_code == 200
    assert upload_response.get_json()["created"] == 1

    from app.extensions import db
    from app.models.enrollment import Enrollment
    from app.models.student import Student
    from app.models.student_subject import StudentSubject

    with app.app_context():
        student = db.session.query(Student).filter_by(registration_number="TVET/2026/100").one()
        enrollment = db.session.query(Enrollment).filter_by(student_id=student.id).one()
        assert enrollment.module_id == ids["module_id"]
        assigned_subjects = {
            row.subject_id
            for row in db.session.query(StudentSubject).filter_by(student_id=student.id).all()
        }
        assert assigned_subjects == {ids["subject_id"]}
        assert ids["other_subject_id"] not in assigned_subjects
